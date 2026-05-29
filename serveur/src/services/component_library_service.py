"""Helpers for importing and exporting the component library Excel format."""

from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha1
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..models.bom import BomItem, Component
from ..services.machine_footprint_catalog_service import MachineFootprintCatalogService
from ..utils.feeder_types import normalize_component_feeder_type


EXPECTED_HEADERS = [
    "Value",
    "MPN",
    "EagleFootprint",
    "MachineFootprint",
    "FeederType",
    "PitchMm",
]

LEGACY_HEADERS = EXPECTED_HEADERS[:-1]


@dataclass
class ComponentLibraryImportResult:
    created_count: int
    updated_count: int
    skipped_count: int
    item_count: int
    errors: List[str]


class ComponentLibraryService:
    """Handle round-trip compatible component library workbooks."""

    machine_footprint_catalog_service = MachineFootprintCatalogService()

    @staticmethod
    def clean_text(value) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @classmethod
    def normalize_footprint(cls, value) -> Optional[str]:
        text = cls.clean_text(value)
        return text.upper() if text else None

    @classmethod
    def normalize_lookup_token(cls, value) -> Optional[str]:
        text = cls.clean_text(value)
        return text.upper() if text else None

    @classmethod
    def normalize_pitch_mm(cls, value) -> Optional[float]:
        if value is None:
            return None

        if isinstance(value, (int, float)):
            return float(value)

        text = cls.clean_text(value)
        if not text:
            return None

        normalized = text.replace(",", ".")
        try:
            return float(normalized)
        except ValueError as exc:
            raise ValueError(f"Invalid PitchMm value: {value}") from exc

    @classmethod
    def build_component_reference(
        cls,
        value: Optional[str],
        mpn: Optional[str],
        footprint_eagle: Optional[str],
    ) -> str:
        key = "|".join(
            [
                (mpn or value or "COMPONENT").strip().upper(),
                (footprint_eagle or "").strip().upper(),
            ]
        )
        digest = sha1(key.encode("utf-8")).hexdigest()[:16].upper()
        return f"LIB-{digest}"

    @classmethod
    def build_lookup(cls, components: Iterable[Component]) -> Dict[Tuple[str, str], Component]:
        """Build a value/MPN + footprint lookup from the component library."""
        lookup: Dict[Tuple[str, str], Component] = {}

        for component in components:
            name_candidates = [component.mpn, component.value, component.description]
            footprint_candidates = [component.footprint_pnp, component.package, component.footprint_eagle]

            for name in name_candidates:
                normalized_name = cls.normalize_lookup_token(name)
                if not normalized_name:
                    continue

                for footprint in footprint_candidates:
                    normalized_footprint = cls.normalize_lookup_token(footprint)
                    if not normalized_footprint:
                        continue

                    lookup.setdefault((normalized_name, normalized_footprint), component)

        return lookup

    @classmethod
    def match_candidates(
        cls,
        lookup: Dict[Tuple[str, str], Component],
        value_candidates: Iterable[Optional[str]],
        footprint_candidates: Iterable[Optional[str]],
    ) -> Optional[Component]:
        """Match a component from candidate names/values and candidate footprints."""
        for value in value_candidates:
            normalized_value = cls.normalize_lookup_token(value)
            if not normalized_value:
                continue

            for footprint in footprint_candidates:
                normalized_footprint = cls.normalize_lookup_token(footprint)
                if not normalized_footprint:
                    continue

                component = lookup.get((normalized_value, normalized_footprint))
                if component:
                    return component

        return None

    @classmethod
    def match_bom_item(
        cls,
        lookup: Dict[Tuple[str, str], Component],
        bom_item: BomItem,
    ) -> Optional[Component]:
        """Match a persisted BOM item against the component library."""
        return cls.match_candidates(
            lookup,
            value_candidates=[bom_item.value_raw, bom_item.value_harmonized],
            footprint_candidates=[bom_item.footprint_pnp, bom_item.footprint_eagle],
        )

    @classmethod
    def match_item_payload(
        cls,
        lookup: Dict[Tuple[str, str], Component],
        item: Dict[str, object],
    ) -> Optional[Component]:
        """Match a serialized/imported BOM item dict against the component library."""
        return cls.match_candidates(
            lookup,
            value_candidates=[item.get("value_raw"), item.get("value_harmonized")],
            footprint_candidates=[item.get("footprint_pnp"), item.get("footprint_eagle")],
        )

    @classmethod
    def component_to_library_row(cls, component: Component) -> Dict[str, object]:
        return {
            "Value": component.value or "",
            "MPN": component.mpn or "",
            "EagleFootprint": component.footprint_eagle or "",
            "MachineFootprint": component.footprint_pnp or component.package or "",
            "FeederType": normalize_component_feeder_type(component.feeder_type) or "",
            "PitchMm": "" if component.pitch_mm is None else component.pitch_mm,
        }

    @classmethod
    def validate_headers(cls, headers: Iterable[object]) -> None:
        normalized_headers = [cls.clean_text(header) or "" for header in headers]
        if normalized_headers[: len(EXPECTED_HEADERS)] not in (EXPECTED_HEADERS, LEGACY_HEADERS):
            raise ValueError(
                "Invalid component library header. Expected: "
                + ", ".join(EXPECTED_HEADERS)
                + " (legacy format without PitchMm is also accepted)"
            )

    @classmethod
    def import_workbook(
        cls,
        file_path: str,
        db: Session,
    ) -> ComponentLibraryImportResult:
        workbook = load_workbook(Path(file_path), data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        headers = [cell.value for cell in worksheet[1]]
        cls.validate_headers(headers)
        has_pitch_column = len(headers) >= len(EXPECTED_HEADERS)

        created_count = 0
        updated_count = 0
        skipped_count = 0
        item_count = 0
        errors: List[str] = []

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True),
            start=2,
        ):
            value = cls.clean_text(row[0])
            mpn = cls.clean_text(row[1])
            footprint_eagle = cls.normalize_footprint(row[2])
            footprint_pnp = cls.clean_text(row[3])
            feeder_type = normalize_component_feeder_type(row[4])
            pitch_mm_value = row[5] if has_pitch_column and len(row) > 5 else None
            pitch_mm = cls.normalize_pitch_mm(pitch_mm_value)

            if not any([value, mpn, footprint_eagle, footprint_pnp, feeder_type, pitch_mm is not None]):
                skipped_count += 1
                continue

            if not value and not mpn:
                errors.append(f"Row {row_index}: missing Value/MPN")
                continue

            reference = cls.build_component_reference(value, mpn, footprint_eagle)
            component = db.query(Component).filter(Component.reference == reference).first()
            is_new = component is None

            if is_new:
                component = Component(reference=reference)
                db.add(component)
                created_count += 1
            else:
                updated_count += 1

            component.value = value
            component.mpn = mpn
            component.footprint_eagle = footprint_eagle
            component.footprint_pnp = footprint_pnp
            component.feeder_type = feeder_type
            component.pitch_mm = pitch_mm
            component.package = footprint_pnp or component.package
            cls.machine_footprint_catalog_service.apply_defaults_to_component(
                db,
                component,
                overwrite=False,
            )

            item_count += 1

        db.commit()

        return ComponentLibraryImportResult(
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            item_count=item_count,
            errors=errors,
        )

    @classmethod
    def export_workbook(cls, components: Iterable[Component]) -> BytesIO:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Component Library"
        worksheet.append(EXPECTED_HEADERS)

        sorted_components = sorted(
            components,
            key=lambda component: (
                (component.value or "").upper(),
                (component.mpn or "").upper(),
                (component.footprint_eagle or "").upper(),
                (component.footprint_pnp or "").upper(),
            ),
        )

        for component in sorted_components:
            row = cls.component_to_library_row(component)
            worksheet.append([row[header] for header in EXPECTED_HEADERS])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
