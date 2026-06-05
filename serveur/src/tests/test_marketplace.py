"""
Tests for Marketplace endpoints: machines, fixed feeders, footprint mappings, productions.
"""
import json
import os
import sys
from io import BytesIO
import tempfile

import pytest
from openpyxl import Workbook, load_workbook

from sqlalchemy.orm import Session
from tests.conftest import client, TestingSessionLocal
from src.models.production import Production
from src.models.bom import BomItem, BomReference, BomRevision

def test_machine_listing_exposes_description_and_counts():
    """Machine list should expose the fields needed by the Machine PnP page."""
    create_machine_response = client.post(
        "/api/marketplace/machines",
        json={
            "name": "PNP-ALPHA",
            "num_positions": 80,
            "description": "Ligne principale",
            "notes": "Equipe A",
        },
    )
    assert create_machine_response.status_code == 200
    machine_id = create_machine_response.json()["machine_id"]

    feeder_response = client.post(
        "/api/marketplace/feeder-types",
        json={
            "size_mm": 8,
            "capacity": 300,
            "description": "8 mm standard",
        },
    )
    assert feeder_response.status_code == 200
    feeder_id = feeder_response.json()["feeder_id"]

    assign_response = client.post(f"/api/marketplace/machines/{machine_id}/feeder-types/{feeder_id}")
    assert assign_response.status_code == 200

    list_response = client.get("/api/marketplace/machines")
    assert list_response.status_code == 200

    machine = next(item for item in list_response.json()["data"] if item["id"] == machine_id)
    assert machine["description"] == "Ligne principale"
    assert machine["notes"] == "Equipe A"
    assert machine["assigned_feeder_types"] == 1
    assert machine["active_production_plans"] == 0


def test_update_machine_can_change_total_positions():
    """Machine update endpoint should allow editing the total feeder positions."""
    create_machine_response = client.post(
        "/api/marketplace/machines",
        json={
            "name": "PNP-BETA",
            "num_positions": 60,
        },
    )
    assert create_machine_response.status_code == 200
    machine_id = create_machine_response.json()["machine_id"]

    update_response = client.put(
        f"/api/marketplace/machines/{machine_id}",
        json={
            "name": "PNP-BETA",
            "num_positions": 80,
            "description": "Machine atelier",
        },
    )
    assert update_response.status_code == 200

    get_response = client.get(f"/api/marketplace/machines/{machine_id}")
    assert get_response.status_code == 200
    assert get_response.json()["num_positions"] == 80
    assert get_response.json()["description"] == "Machine atelier"


def test_production_can_be_assigned_to_machine_and_seen_in_machine_summary():
    """A production assigned to a machine should appear in the machine detail summary."""
    create_machine_response = client.post(
        "/api/marketplace/machines",
        json={
            "name": "PNP-GAMMA",
            "num_positions": 80,
        },
    )
    assert create_machine_response.status_code == 200
    machine_id = create_machine_response.json()["machine_id"]

    create_production_response = client.post(
        "/api/marketplace/productions",
        json={
            "name": "prod-machine-01",
        },
    )
    assert create_production_response.status_code == 200
    production_id = create_production_response.json()["id"]

    assign_response = client.patch(
        f"/api/marketplace/productions/{production_id}",
        json={
            "machine_id": machine_id,
        },
    )
    assert assign_response.status_code == 200
    assert assign_response.json()["machine_id"] == machine_id
    assert assign_response.json()["machine_name"] == "PNP-GAMMA"

    summary_response = client.get(f"/api/marketplace/machines/{machine_id}/summary")
    assert summary_response.status_code == 200
    assert summary_response.json()["assigned_productions"] == 1
    assert summary_response.json()["productions"][0]["id"] == production_id


def test_machine_production_order_validation_and_feeder_plan_endpoint():
    """A machine can persist the BOM order of a production, validate it, and expose the computed feeder plan."""
    create_machine_response = client.post(
        "/api/marketplace/machines",
        json={
            "name": "PNP-PLAN-API",
            "num_positions": 6,
        },
    )
    assert create_machine_response.status_code == 200
    machine_id = create_machine_response.json()["machine_id"]

    feeder_8_response = client.post(
        "/api/marketplace/feeder-types",
        json={
            "size_mm": 8,
            "capacity": 200,
        },
    )
    assert feeder_8_response.status_code == 200
    feeder_8_id = feeder_8_response.json()["feeder_id"]

    feeder_12_response = client.post(
        "/api/marketplace/feeder-types",
        json={
            "size_mm": 12,
            "capacity": 120,
        },
    )
    assert feeder_12_response.status_code == 200
    feeder_12_id = feeder_12_response.json()["feeder_id"]

    assert client.post(f"/api/marketplace/machines/{machine_id}/feeder-types/{feeder_8_id}").status_code == 200
    assert client.post(f"/api/marketplace/machines/{machine_id}/feeder-types/{feeder_12_id}").status_code == 200

    cart_response = client.post(
        "/api/marketplace/carts",
        json={
            "name": "COMPOSANT_COMMUN_API",
            "kind": "COMMON",
            "capacity_positions": 20,
        },
    )
    assert cart_response.status_code == 200
    common_cart_id = cart_response.json()["cart_id"]

    fixed_component_response = client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-10K-PLAN-API",
            "value": "10K",
            "footprint_eagle": "RES_0603",
            "footprint_pnp": "RES_0603",
            "feeder_type": "8mm",
            "is_fixed_feeder": True,
            "fixed_cart_id": common_cart_id,
        },
    )
    assert fixed_component_response.status_code == 200

    opamp_component_response = client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-OPA-PLAN-API",
            "value": "OPA828",
            "footprint_eagle": "SOIC-8",
            "footprint_pnp": "SOIC-8",
            "feeder_type": "12mm",
        },
    )
    assert opamp_component_response.status_code == 200

    cap_component_response = client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-100N-PLAN-API",
            "value": "100n",
            "footprint_eagle": "CAP_0603",
            "footprint_pnp": "CAP_0603",
            "feeder_type": "8mm",
        },
    )
    assert cap_component_response.status_code == 200

    db = TestingSessionLocal()
    revision_a_id = None
    revision_b_id = None
    try:
        bom_ref_a = BomReference(reference="BOARD-A-API", category="AMPLI")
        bom_ref_b = BomReference(reference="BOARD-B-API", category="AMPLI")
        db.add_all([bom_ref_a, bom_ref_b])
        db.commit()
        db.refresh(bom_ref_a)
        db.refresh(bom_ref_b)

        revision_a = BomRevision(
            bom_ref_id=bom_ref_a.id,
            revision="REV_A",
            type=BomRevision.TypeEnum.TOP,
            status=BomRevision.StatusEnum.ACTIVE,
        )
        revision_b = BomRevision(
            bom_ref_id=bom_ref_b.id,
            revision="REV_A",
            type=BomRevision.TypeEnum.TOP,
            status=BomRevision.StatusEnum.ACTIVE,
        )
        db.add_all([revision_a, revision_b])
        db.commit()
        db.refresh(revision_a)
        db.refresh(revision_b)
        revision_a_id = revision_a.id
        revision_b_id = revision_b.id

        db.add_all(
            [
                BomItem(
                    bom_revision_id=revision_a.id,
                    reference_item="R1",
                    value_raw="10K",
                    value_harmonized="10K",
                    footprint_eagle="RES_0603",
                    footprint_pnp="RES_0603",
                    quantity=2,
                ),
                BomItem(
                    bom_revision_id=revision_a.id,
                    reference_item="U1",
                    value_raw="OPA828",
                    value_harmonized="OPA828",
                    footprint_eagle="SOIC-8",
                    footprint_pnp="SOIC-8",
                    quantity=1,
                ),
                BomItem(
                    bom_revision_id=revision_b.id,
                    reference_item="R7",
                    value_raw="10K",
                    value_harmonized="10K",
                    footprint_eagle="RES_0603",
                    footprint_pnp="RES_0603",
                    quantity=3,
                ),
                BomItem(
                    bom_revision_id=revision_b.id,
                    reference_item="C4",
                    value_raw="100n",
                    value_harmonized="100n",
                    footprint_eagle="CAP_0603",
                    footprint_pnp="CAP_0603",
                    quantity=2,
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    create_production_response = client.post(
        "/api/marketplace/productions",
        json={
            "name": "prod-plan-api DATE:03/2026",
        },
    )
    assert create_production_response.status_code == 200
    production_id = create_production_response.json()["id"]

    attach_response = client.post(
        f"/api/marketplace/productions/{production_id}/bom-revisions",
        json={"bom_revision_ids": [revision_a_id, revision_b_id]},
    )
    assert attach_response.status_code == 200

    assign_response = client.patch(
        f"/api/marketplace/productions/{production_id}",
        json={"machine_id": machine_id},
    )
    assert assign_response.status_code == 200

    quantity_response = client.patch(
        f"/api/marketplace/productions/{production_id}/bom-quantities",
        json={
            "items": [
                {"bom_revision_id": revision_a_id, "quantity_to_produce": 2},
                {"bom_revision_id": revision_b_id, "quantity_to_produce": 3},
            ],
        },
    )
    assert quantity_response.status_code == 200

    reorder_response = client.patch(
        f"/api/marketplace/machines/{machine_id}/productions/{production_id}/bom-order",
        json={"bom_revision_ids": [revision_b_id, revision_a_id]},
    )
    assert reorder_response.status_code == 200
    ordered_revisions = reorder_response.json()["production"]["bom_revisions"]
    assert [entry["bom_revision_id"] for entry in ordered_revisions] == [revision_b_id, revision_a_id]
    assert reorder_response.json()["production"]["has_validated_order"] is False

    validate_response = client.post(
        f"/api/marketplace/machines/{machine_id}/productions/{production_id}/validate-order",
    )
    assert validate_response.status_code == 200
    validate_payload = validate_response.json()
    assert validate_payload["production"]["has_validated_order"] is True
    assert validate_payload["plan"]["is_order_validated"] is True
    assert [entry["bom_revision_id"] for entry in validate_payload["plan"]["ordered_boms"]] == [revision_b_id, revision_a_id]
    assert [entry["quantity_to_produce"] for entry in validate_payload["plan"]["ordered_boms"]] == [3, 2]
    assert validate_payload["plan"]["quantity_source"] == "PRODUCTION"
    assert validate_payload["plan"]["total_build_quantity"] == 5
    assert validate_payload["plan"]["assigned_fixed_component_count"] == 1
    assert validate_payload["plan"]["assigned_component_count"] == 3
    assert validate_payload["plan"]["slot_assignments"][0]["component_reference"] == "LIB-10K-PLAN-API"
    assert validate_payload["plan"]["slot_assignments"][0]["is_stable_between_boms"] is True
    assert validate_payload["plan"]["slot_assignments"][0]["bom_revision_ids"] == [revision_b_id, revision_a_id]
    assert validate_payload["plan"]["slot_assignments"][0]["total_board_quantity_by_revision"][str(revision_b_id)] == 9
    assert validate_payload["plan"]["slot_assignments"][0]["total_board_quantity_by_revision"][str(revision_a_id)] == 4
    assert validate_payload["plan"]["slot_assignments"][0]["board_quantity_by_revision"][str(revision_b_id)] == 3
    assert validate_payload["plan"]["slot_assignments"][0]["board_quantity_by_revision"][str(revision_a_id)] == 2
    assert validate_payload["plan"]["slot_assignments"][0]["total_board_quantity"] == 13
    assert validate_payload["plan"]["slot_assignments"][0]["average_board_quantity"] == 2.6
    assert validate_payload["plan"]["stable_assignment_count"] == 1
    assert validate_payload["plan"]["bom_assignment_summaries"][0]["assignment_indexes"] == [1, 2]

    plan_response = client.get(
        f"/api/marketplace/machines/{machine_id}/productions/{production_id}/feeder-plan",
    )
    assert plan_response.status_code == 200
    plan_payload = plan_response.json()
    assert plan_payload["is_order_validated"] is True
    assert [entry["bom_revision_id"] for entry in plan_payload["ordered_boms"]] == [revision_b_id, revision_a_id]
    assert plan_payload["slot_assignments"][0]["component_reference"] == "LIB-10K-PLAN-API"
    assert plan_payload["slot_assignments"][1]["total_board_quantity"] == 6
    assert plan_payload["slot_assignments"][2]["total_board_quantity"] == 2
    assert plan_payload["slot_assignments"][2]["is_stable_between_boms"] is False
    assert plan_payload["bom_assignment_summaries"][1]["assignment_indexes"] == [1, 3]


def test_calculate_fixed_feeders_endpoint_updates_component_assignments():
    """Marketplace endpoint should calculate fixed feeders and write cart assignments to the library."""
    common_cart_response = client.post(
        "/api/marketplace/carts",
        json={
            "name": "COMPOSANT_RECURRENT",
            "kind": "COMMON",
            "capacity_positions": 80,
        },
    )
    assert common_cart_response.status_code == 200

    category_cart_response = client.post(
        "/api/marketplace/carts",
        json={
            "name": "AMPLI",
            "kind": "CATEGORY",
            "target_category": "AMPLI",
            "capacity_positions": 40,
        },
    )
    assert category_cart_response.status_code == 200

    resistor_response = client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-10K-0603",
            "value": "10K",
            "footprint_eagle": "RES_0603",
            "footprint_pnp": "RES_0603",
            "feeder_type": "8mm",
        },
    )
    assert resistor_response.status_code == 200

    opamp_response = client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-OPA828",
            "value": "OPA828IDGNT",
            "mpn": "OPA828IDGNT",
            "footprint_eagle": "SOIC-8",
            "footprint_pnp": "SOIC-8",
            "feeder_type": "12mm",
        },
    )
    assert opamp_response.status_code == 200

    db = TestingSessionLocal()
    try:
        amp_ref = BomReference(reference="AMP-API-001", category="AMPLI")
        sensor_ref = BomReference(reference="SNS-API-001", category="SENSOR")
        db.add_all([amp_ref, sensor_ref])
        db.commit()
        db.refresh(amp_ref)
        db.refresh(sensor_ref)

        amp_rev = BomRevision(
            bom_ref_id=amp_ref.id,
            revision="REV_A",
            type=BomRevision.TypeEnum.TOP,
            status=BomRevision.StatusEnum.ACTIVE,
        )
        sensor_rev = BomRevision(
            bom_ref_id=sensor_ref.id,
            revision="REV_A",
            type=BomRevision.TypeEnum.TOP,
            status=BomRevision.StatusEnum.ACTIVE,
        )
        db.add_all([amp_rev, sensor_rev])
        db.commit()
        db.refresh(amp_rev)
        db.refresh(sensor_rev)

        db.add_all(
            [
                BomItem(
                    bom_revision_id=amp_rev.id,
                    reference_item="R1",
                    value_raw="10K",
                    value_harmonized="10K",
                    footprint_eagle="RES_0603",
                    footprint_pnp="RES_0603",
                ),
                BomItem(
                    bom_revision_id=amp_rev.id,
                    reference_item="U1",
                    value_raw="OPA828IDGNT",
                    value_harmonized="OPA828IDGNT",
                    footprint_eagle="SOIC-8",
                    footprint_pnp="SOIC-8",
                ),
                BomItem(
                    bom_revision_id=sensor_rev.id,
                    reference_item="R4",
                    value_raw="10K",
                    value_harmonized="10K",
                    footprint_eagle="RES_0603",
                    footprint_pnp="RES_0603",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    calculate_response = client.post("/api/marketplace/fixed-feeders/calculate")

    assert calculate_response.status_code == 200
    payload = calculate_response.json()
    assert payload["assigned_count"] == 2
    assert payload["assigned_common_count"] == 1
    assert payload["assigned_category_count"] == 1

    component_list_response = client.get("/api/bom/components")
    assert component_list_response.status_code == 200
    components_by_reference = {
        component["reference"]: component
        for component in component_list_response.json()
    }
    assert components_by_reference["LIB-10K-0603"]["fixed_cart_name"] == "COMPOSANT_RECURRENT"
    assert components_by_reference["LIB-OPA828"]["fixed_cart_name"] == "AMPLI"


def test_calculate_fixed_feeders_endpoint_ignores_draft_boms():
    """Draft imports should not influence fixed-feeder calculation until the BOM is validated."""
    cart_response = client.post(
        "/api/marketplace/carts",
        json={
            "name": "COMPOSANT_RECURRENT",
            "kind": "COMMON",
            "capacity_positions": 80,
        },
    )
    assert cart_response.status_code == 200

    component_response = client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-10K-0603",
            "value": "10K",
            "footprint_eagle": "RES_0603",
            "footprint_pnp": "RES_0603",
            "feeder_type": "8mm",
        },
    )
    assert component_response.status_code == 200

    db = TestingSessionLocal()
    try:
        draft_ref = BomReference(reference="DRAFT-ONLY-001", category="AMPLI")
        db.add(draft_ref)
        db.commit()
        db.refresh(draft_ref)

        draft_rev = BomRevision(
            bom_ref_id=draft_ref.id,
            revision="REV_A",
            type=BomRevision.TypeEnum.TOP,
            status=BomRevision.StatusEnum.DRAFT,
        )
        db.add(draft_rev)
        db.commit()
        db.refresh(draft_rev)

        db.add(
            BomItem(
                bom_revision_id=draft_rev.id,
                reference_item="R1",
                value_raw="10K",
                value_harmonized="10K",
                footprint_eagle="RES_0603",
                footprint_pnp="RES_0603",
            )
        )
        db.commit()
    finally:
        db.close()

    calculate_response = client.post("/api/marketplace/fixed-feeders/calculate")
    assert calculate_response.status_code == 400
    assert "BOM ACTIVE" in calculate_response.json()["detail"]


def test_list_fixed_feeders_endpoint_returns_usage_statistics():
    """Fixed-feeder listing should expose BOM overlap and average quantity per board."""
    cart_response = client.post(
        "/api/marketplace/carts",
        json={
            "name": "COMPOSANT_RECURRENT",
            "kind": "COMMON",
            "capacity_positions": 80,
        },
    )
    assert cart_response.status_code == 200
    cart_id = cart_response.json()["cart_id"]

    component_response = client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-22K-0603",
            "value": "22K",
            "footprint_eagle": "RES_0603",
            "footprint_pnp": "RES_0603",
            "feeder_type": "8mm",
            "is_fixed_feeder": True,
            "fixed_cart_id": cart_id,
        },
    )
    assert component_response.status_code == 200

    db = TestingSessionLocal()
    try:
        amp_ref = BomReference(reference="AMP-LIST-001", category="AMPLI")
        amp_ref_2 = BomReference(reference="AMP-LIST-002", category="AMPLI")
        db.add_all([amp_ref, amp_ref_2])
        db.commit()
        db.refresh(amp_ref)
        db.refresh(amp_ref_2)

        rev_a = BomRevision(
            bom_ref_id=amp_ref.id,
            revision="REV_A",
            type=BomRevision.TypeEnum.TOP,
            status=BomRevision.StatusEnum.ACTIVE,
        )
        rev_b = BomRevision(
            bom_ref_id=amp_ref_2.id,
            revision="REV_A",
            type=BomRevision.TypeEnum.TOP,
            status=BomRevision.StatusEnum.ACTIVE,
        )
        db.add_all([rev_a, rev_b])
        db.commit()
        db.refresh(rev_a)
        db.refresh(rev_b)

        db.add_all(
            [
                BomItem(
                    bom_revision_id=rev_a.id,
                    reference_item="R1",
                    value_raw="22K",
                    value_harmonized="22K",
                    footprint_eagle="RES_0603",
                    footprint_pnp="RES_0603",
                    quantity=2,
                ),
                BomItem(
                    bom_revision_id=rev_b.id,
                    reference_item="R9",
                    value_raw="22K",
                    value_harmonized="22K",
                    footprint_eagle="RES_0603",
                    footprint_pnp="RES_0603",
                    quantity=4,
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/api/marketplace/fixed-feeders/components")

    assert response.status_code == 200
    payload = response.json()
    row = next(item for item in payload["data"] if item["reference"] == "LIB-22K-0603")
    assert row["bom_reference_count"] == 2
    assert row["average_board_quantity"] == 3.0
    assert row["feeder_size_mm"] == 8
    assert row["fixed_cart_name"] == "COMPOSANT_RECURRENT"


def test_patch_fixed_feeder_component_endpoint_allows_manual_assignment():
    """A component should be manually assignable as a fixed feeder from the marketplace UI."""
    cart_response = client.post(
        "/api/marketplace/carts",
        json={
            "name": "MANUEL-FIXE",
            "kind": "CUSTOM",
            "capacity_positions": 40,
        },
    )
    assert cart_response.status_code == 200
    cart_id = cart_response.json()["cart_id"]

    feeder_response = client.post(
        "/api/marketplace/feeder-types",
        json={
            "size_mm": 16,
            "capacity": 120,
            "description": "16 mm",
        },
    )
    assert feeder_response.status_code == 200
    feeder_id = feeder_response.json()["feeder_id"]

    component_response = client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-CONN-USB",
            "value": "USB-C",
            "footprint_eagle": "USB_C",
            "footprint_pnp": "USB_C",
        },
    )
    assert component_response.status_code == 200
    component_id = component_response.json()["id"]

    patch_response = client.patch(
        f"/api/marketplace/fixed-feeders/components/{component_id}",
        json={
            "is_fixed_feeder": True,
            "fixed_cart_id": cart_id,
            "feeder_id": feeder_id,
        },
    )

    assert patch_response.status_code == 200
    component_detail_response = client.get(f"/api/bom/components/{component_id}")
    assert component_detail_response.status_code == 200
    data = component_detail_response.json()
    assert data["is_fixed_feeder"] is True
    assert data["fixed_cart_id"] == cart_id
    assert data["fixed_cart_name"] == "MANUEL-FIXE"
    assert data["feeder_type"] == "CL16"


@pytest.mark.parametrize(
    "path",
    [
        "/api/marketplace/machines/999",
        "/api/marketplace/carts/999",
        "/api/marketplace/feeder-types/999",
    ],
)
def test_marketplace_delete_missing_resources_return_404(path):
    """Delete endpoints should preserve 404 responses for missing entities."""
    response = client.delete(path)

    assert response.status_code == 404


def test_update_component_refreshes_footprint_mapping_for_future_imports():
    """Editing the component library should refresh the reusable Eagle -> PnP mapping."""
    client.post(
        "/api/bom/mappings/footprints",
        json={
            "footprint_eagle": "SOP65P490X110-9N",
            "footprint_pnp": "SOP65P490X110-9N",
        },
    )

    create_response = client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-OPA828",
            "value": "OPA828IDGNT",
            "mpn": "OPA828IDGNT",
            "package": "SOP65P490X110-9N",
            "footprint_eagle": "SOP65P490X110-9N",
            "footprint_pnp": "SOP65P490X110-9N",
        },
    )
    component_id = create_response.json()["id"]

    update_response = client.put(
        f"/api/bom/components/{component_id}",
        json={
            "id": component_id,
            "reference": "LIB-OPA828",
            "value": "OPA828IDGNT",
            "mpn": "OPA828IDGNT",
            "package": "SOIC-8",
            "supplier_code": None,
            "footprint_eagle": "SOP65P490X110-9N",
            "footprint_pnp": "SOIC-8",
            "feeder_type": None,
            "description": None,
            "notes": "Updated from settings",
        },
    )

    assert update_response.status_code == 200
    mappings_response = client.get("/api/bom/mappings/footprints?search=SOP65P490X110-9N")
    assert mappings_response.status_code == 200
    mappings = mappings_response.json()
    assert len(mappings) == 1
    assert mappings[0]["footprint_pnp"] == "SOIC-8"

    bom_content = """Reference Value Footprint X Y Rotation Type
IC2 OPA828IDGNT SOP65P490X110-9N 10.0 20.0 0 IC
"""

    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as f:
        f.write(bom_content)
        temp_path = f.name

    try:
        with open(temp_path, 'rb') as f:
            files = {'file': ('opa828_updated_bom.txt', f, 'text/plain')}
            import_response = client.post(
                "/api/bom/import",
                files=files,
                params={"reference": "OPA_CARD_UPDATED", "revision": "REV_A", "side": "TOP"},
            )

        assert import_response.status_code == 200
        assert import_response.json()["items"][0]["footprint_pnp"] == "SOIC-8"
    finally:
        os.unlink(temp_path)


def test_get_nonexistent_component():
    """Test getting a component that doesn't exist"""
    response = client.get("/api/bom/components/99999")
    assert response.status_code == 404


def test_import_component_library_workbook():
    """Test importing a component library workbook and persisting enriched fields."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Component Library"
    worksheet.append(["Value", "MPN", "EagleFootprint", "MachineFootprint", "FeederType"])
    worksheet.append(["10nF", "", "CAPC1608X90N", "0603", "8mm"])
    worksheet.append(["0679H5000-01", "0154.500DR", "0154.250DR", "FUSE-SMD", "12mm"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/bom/components/library/import",
        files={
            "file": (
                "component_library.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["item_count"] == 2
    assert data["created_count"] == 2

    list_response = client.get("/api/bom/components")
    assert list_response.status_code == 200
    components = list_response.json()
    assert len(components) == 2
    first_match = next(component for component in components if component["value"] == "10nF")
    second_match = next(component for component in components if component["mpn"] == "0154.500DR")
    assert first_match["footprint_eagle"] == "CAPC1608X90N"
    assert first_match["footprint_pnp"] == "0603"
    assert first_match["feeder_type"] == "CL8-4"
    assert first_match["pitch_mm"] is None
    assert second_match["package"] == "FUSE-SMD"
    assert second_match["pitch_mm"] is None


def test_import_component_library_workbook_reports_partial_errors():
    """Import should report row-level problems without hiding successfully imported rows."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Component Library"
    worksheet.append(["Value", "MPN", "EagleFootprint", "MachineFootprint", "FeederType", "PitchMm"])
    worksheet.append(["10nF", "", "CAPC1608X90N", "0603", "8mm", 0.8])
    worksheet.append(["", "", "CAPC3216X180N", "1206", "8mm", None])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/bom/components/library/import",
        files={
            "file": (
                "component_library_partial.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["success"] is False
    assert data["item_count"] == 1
    assert data["created_count"] == 1
    assert "Row 3: missing Value/MPN" in data["errors"]


def test_import_component_library_workbook_with_pitch_mm():
    """Test importing the extended component library format with PitchMm."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Component Library"
    worksheet.append(["Value", "MPN", "EagleFootprint", "MachineFootprint", "FeederType", "PitchMm"])
    worksheet.append(["10nF", "", "CAPC1608X90N", "0603", "8mm", 0.8])
    worksheet.append(["0679H5000-01", "0154.500DR", "0154.250DR", "FUSE-SMD", "12mm", "1.25"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/bom/components/library/import",
        files={
            "file": (
                "component_library_pitch.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["item_count"] == 2

    list_response = client.get("/api/bom/components")
    assert list_response.status_code == 200
    components = list_response.json()
    first_match = next(component for component in components if component["value"] == "10nF")
    second_match = next(component for component in components if component["mpn"] == "0154.500DR")
    assert first_match["pitch_mm"] == 0.8
    assert second_match["pitch_mm"] == 1.25


def test_export_component_library_workbook():
    """Test exporting the component library in the expected Excel format."""
    client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-TEST-1",
            "value": "10nF",
            "mpn": None,
            "package": "0603",
            "footprint_eagle": "CAPC1608X90N",
            "footprint_pnp": "0603",
            "feeder_type": "8mm",
            "pitch_mm": 0.8,
        },
    )

    response = client.get("/api/bom/components/library/export")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook["Component Library"]
    assert worksheet.cell(1, 1).value == "Value"
    assert worksheet.cell(1, 2).value == "MPN"
    assert worksheet.cell(1, 3).value == "EagleFootprint"
    assert worksheet.cell(1, 4).value == "MachineFootprint"
    assert worksheet.cell(1, 5).value == "FeederType"
    assert worksheet.cell(1, 6).value == "PitchMm"
    assert worksheet.cell(2, 1).value == "10nF"
    assert worksheet.cell(2, 3).value == "CAPC1608X90N"
    assert worksheet.cell(2, 4).value == "0603"
    assert worksheet.cell(2, 5).value == "CL8-4"
    assert worksheet.cell(2, 6).value == 0.8


def test_export_command_erp_workbook():
    """Test ERP export for a generated command."""
    client.post(
        "/api/bom/components",
        json={
            "reference": "LIB-ERP-1",
            "value": "10R",
            "mpn": "RES-10R",
            "package": "0805",
            "supplier_code": "SUP-10R",
            "footprint_eagle": "0805",
            "footprint_pnp": "0805",
            "feeder_type": "8mm",
        },
    )

    bom_content = """Reference Value Footprint X Y Rotation Type
R1 10R 0805 10.0 20.0 0 R
"""

    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as f:
        f.write(bom_content)
        temp_path = f.name

    try:
        with open(temp_path, 'rb') as f:
            files = {'file': ('erp_bom.txt', f, 'text/plain')}
            import_response = client.post(
                "/api/bom/import",
                files=files,
                params={"reference": "ERP_CARD", "revision": "REV_A", "side": "TOP"},
            )

        assert import_response.status_code == 200
        import_data = import_response.json()
        bom_revision_id = import_data["bom_revision_id"]

        create_command_response = client.post(
            "/api/marketplace/commands",
            json={"name": "CMD-ERP-001", "notes": "ERP export test"},
        )
        assert create_command_response.status_code == 200
        command_id = create_command_response.json()["id"]

        add_item_response = client.post(
            f"/api/marketplace/commands/{command_id}/items",
            json={"bom_revision_id": bom_revision_id, "quantity": 3},
        )
        assert add_item_response.status_code == 200

        export_response = client.post(
            f"/api/marketplace/commands/{command_id}/erp-export",
            json={
                "project": "PJ2601-00241 - Achat projet client 2026",
                "delay": "URGENT",
                "remark": "mise en bobine",
                "validator": "Kevin Surrier",
                "requester": "Eric Bouquet",
                "unit": "pièce",
                "default_supplier": "MOUSER",
            },
        )

        assert export_response.status_code == 200
        assert export_response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        workbook = load_workbook(BytesIO(export_response.content))
        worksheet = workbook["Purchase List ERP"]
        # New 12-column ERP layout (audit 2026-06-03 §6.2).
        assert worksheet.cell(1, 1).value == "Référence fournisseur"
        assert worksheet.cell(1, 2).value == "Fournisseur"
        assert worksheet.cell(1, 3).value == "Description"
        assert worksheet.cell(1, 4).value == "Lien web"
        assert worksheet.cell(1, 5).value == "Référence KT"
        assert worksheet.cell(1, 6).value == "Quantité"
        assert worksheet.cell(1, 7).value == "Unité"
        assert worksheet.cell(1, 8).value == "Projet"
        assert worksheet.cell(1, 9).value == "Demandeur"
        assert worksheet.cell(1, 10).value == "Validateur"
        assert worksheet.cell(1, 11).value == "Délai"
        assert worksheet.cell(1, 12).value == "Remarques"
        # No supplier offer cached -> falls back to component library + default supplier.
        assert worksheet.cell(2, 1).value == "SUP-10R"           # supplier_code
        assert worksheet.cell(2, 2).value == "MOUSER"            # default supplier
        assert "RES-10R" in worksheet.cell(2, 3).value           # description from MPN
        assert worksheet.cell(2, 5).value == "LIB-ERP-1"         # Référence KT = COMPONENTS.reference
        assert worksheet.cell(2, 6).value == 3                   # quantity
        assert worksheet.cell(2, 7).value == "pièce"             # unit
        assert worksheet.cell(2, 8).value == "PJ2601-00241 - Achat projet client 2026"
        assert worksheet.cell(2, 9).value == "Eric Bouquet"      # requester
        assert worksheet.cell(2, 10).value == "Kevin Surrier"    # validator
        assert worksheet.cell(2, 11).value == "URGENT"           # delay
        assert worksheet.cell(2, 12).value == "mise en bobine"   # remark
    finally:
        os.unlink(temp_path)


def test_generate_command_batch_returns_summary():
    """Test batch command generation endpoint returns the final summary directly."""
    bom_content = """Reference Value Footprint X Y Rotation Type
R1 10R 0805 10.0 20.0 0 R
"""

    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as f:
        f.write(bom_content)
        temp_path = f.name

    try:
        with open(temp_path, 'rb') as source:
            import_response = client.post(
                "/api/bom/import",
                files={'file': ('cmd_batch.txt', source, 'text/plain')},
                params={"reference": "CMD_BATCH_CARD", "revision": "REV_A", "side": "TOP"},
            )

        assert import_response.status_code == 200
        bom_revision_id = import_response.json()["bom_revision_id"]

        generate_response = client.post(
            "/api/marketplace/commands/generate",
            json={
                "name": "CMD-BATCH-API",
                "notes": "Generated in one call",
                "items": [
                    {
                        "bom_revision_id": bom_revision_id,
                        "quantity": 6,
                    }
                ],
            },
        )

        assert generate_response.status_code == 200
        summary = generate_response.json()
        assert summary["name"] == "CMD-BATCH-API"
        assert summary["items_count"] == 1
        assert summary["total_boards_to_produce"] == 6
        assert len(summary["aggregated_components"]) == 1
        assert summary["aggregated_components"][0]["quantity"] == 6
    finally:
        os.unlink(temp_path)


def test_delete_command_returns_404_when_missing():
    """Deleting a missing command should preserve the 404 response."""
    response = client.delete("/api/marketplace/commands/999")

    assert response.status_code == 404
    assert response.json()["detail"] == "Command 999 not found"


def test_update_command_returns_404_when_missing():
    """Updating a missing command should preserve the 404 response."""
    response = client.put(
        "/api/marketplace/commands/999",
        json={"name": "CMD-MISSING"},
    )

    assert response.status_code == 404
    assert response.json()["detail"] == "Command 999 not found"


def test_reports_top_components_aggregate_by_logical_component():
    """The reporting endpoint should aggregate by logical component, not by BOM designator."""
    db = TestingSessionLocal()
    try:
        bom_ref = BomReference(reference="REPORT_CARD")
        db.add(bom_ref)
        db.commit()
        db.refresh(bom_ref)

        revision = BomRevision(
            bom_ref_id=bom_ref.id,
            revision="REV_A",
            type=BomRevision.TypeEnum.TOP,
            status=BomRevision.StatusEnum.ACTIVE,
        )
        db.add(revision)
        db.commit()
        db.refresh(revision)

        db.add_all(
            [
                BomItem(
                    bom_revision_id=revision.id,
                    reference_item="R1",
                    value_raw="10K",
                    value_harmonized="10K",
                    footprint_eagle="RES_0603",
                    footprint_pnp="RES_0603",
                    component_type="R",
                    quantity=2,
                ),
                BomItem(
                    bom_revision_id=revision.id,
                    reference_item="R2",
                    value_raw="10K",
                    value_harmonized="10K",
                    footprint_eagle="RES_0603",
                    footprint_pnp="RES_0603",
                    component_type="R",
                    quantity=3,
                ),
                BomItem(
                    bom_revision_id=revision.id,
                    reference_item="R3",
                    value_raw="10K",
                    value_harmonized="10K",
                    footprint_eagle="RES_0603",
                    footprint_pnp="RES_0603",
                    component_type="R",
                    quantity=4,
                    dnp=True,
                ),
            ]
        )

        command = Command(name="CMD-REPORT-001")
        db.add(command)
        db.commit()
        db.refresh(command)

        db.add(
            CommandItem(
                command_id=command.id,
                bom_revision_id=revision.id,
                quantity_to_produce=2,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get("/api/reports/components/top")

    assert response.status_code == 200
    rows = response.json()
    assert len(rows) == 1
    assert rows[0]["value"] == "10K"
    assert rows[0]["footprint"] == "RES_0603"
    assert rows[0]["component_type"] == "R"
    assert rows[0]["total_required"] == 10


def test_list_productions_does_not_mutate_statuses():
    """Listing productions must stay read-only and not auto-activate a production."""
    db = TestingSessionLocal()
    try:
        production = Production(
            name="prod-read-only",
            status=Production.StatusEnum.DRAFT,
        )
        db.add(production)
        db.commit()
        production_id = production.id
    finally:
        db.close()

    response = client.get("/api/marketplace/productions")

    assert response.status_code == 200
    item = next(entry for entry in response.json()["items"] if entry["id"] == production_id)
    assert item["status"] == "DRAFT"

    db = TestingSessionLocal()
    try:
        stored = db.query(Production).filter(Production.id == production_id).first()
        assert stored is not None
        assert stored.status == Production.StatusEnum.DRAFT
    finally:
        db.close()


def test_create_footprint_mapping():
    """Test creating and listing a footprint mapping."""
    response = client.post(
        "/api/bom/mappings/footprints",
        json={
            "footprint_eagle": "resc1608x55n",
            "footprint_pnp": "R_0603",
            "machine_compatible": "PNP-01",
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["footprint_eagle"] == "RESC1608X55N"
    assert data["footprint_pnp"] == "R_0603"

    list_response = client.get("/api/bom/mappings/footprints")
    assert list_response.status_code == 200
    mappings = list_response.json()
    assert len(mappings) == 1
    assert mappings[0]["footprint_eagle"] == "RESC1608X55N"


def test_update_footprint_mapping():
    """Test updating an existing Eagle -> PnP mapping."""
    create_response = client.post(
        "/api/bom/mappings/footprints",
        json={
            "footprint_eagle": "resc1608x55n",
            "footprint_pnp": "R_0603",
            "machine_compatible": "PNP-01",
        },
    )
    mapping_id = create_response.json()["id"]

    update_response = client.put(
        f"/api/bom/mappings/footprints/{mapping_id}",
        json={
            "footprint_eagle": "resc1608x55n",
            "footprint_pnp": "R_0603_PRECISION",
            "machine_compatible": "PNP-02",
            "notes": "Mis a jour depuis Parametre",
        },
    )

    assert update_response.status_code == 200
    data = update_response.json()
    assert data["footprint_eagle"] == "RESC1608X55N"
    assert data["footprint_pnp"] == "R_0603_PRECISION"
    assert data["machine_compatible"] == "PNP-02"
    assert data["notes"] == "Mis a jour depuis Parametre"


def test_production_can_be_deleted():
    """DELETE /productions/{id} removes the production and 404s afterwards."""
    create_response = client.post(
        "/api/marketplace/productions",
        json={"name": "prod-to-delete"},
    )
    assert create_response.status_code == 200
    production_id = create_response.json()["id"]

    delete_response = client.delete(f"/api/marketplace/productions/{production_id}")
    assert delete_response.status_code == 200
    assert delete_response.json()["status"] == "deleted"

    get_response = client.get(f"/api/marketplace/productions/{production_id}")
    assert get_response.status_code == 404

    second_delete = client.delete(f"/api/marketplace/productions/{production_id}")
    assert second_delete.status_code == 404


def test_production_can_be_duplicated():
    """POST /productions/{id}/duplicate creates a copy under a unique name."""
    create_response = client.post(
        "/api/marketplace/productions",
        json={"name": "prod-source"},
    )
    assert create_response.status_code == 200
    source_id = create_response.json()["id"]

    duplicate_response = client.post(f"/api/marketplace/productions/{source_id}/duplicate")
    assert duplicate_response.status_code == 200
    first_copy = duplicate_response.json()
    assert first_copy["id"] != source_id
    assert first_copy["name"] == "Copie de prod-source"
    assert first_copy["status"] == "DRAFT"

    second_duplicate = client.post(f"/api/marketplace/productions/{source_id}/duplicate")
    assert second_duplicate.status_code == 200
    assert second_duplicate.json()["name"] == "Copie de prod-source (2)"

    # Cleanup to limit cross-test pollution in the shared test DB.
    for pid in (first_copy["id"], second_duplicate.json()["id"], source_id):
        client.delete(f"/api/marketplace/productions/{pid}")

